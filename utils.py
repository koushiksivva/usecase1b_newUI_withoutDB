import os
import io
import re
import tempfile
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
import hashlib
import time
from datetime import datetime
import json

import fitz  # PyMuPDF
from fuzzywuzzy import fuzz
from task_batches import task_batches 
from dotenv import load_dotenv

# Langchain + Azure OpenAI
from langchain_openai import AzureChatOpenAI, AzureOpenAIEmbeddings
from langchain_core.prompts import PromptTemplate
from langchain_core.messages import HumanMessage
from langchain.text_splitter import RecursiveCharacterTextSplitter 

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import base64
from openpyxl.styles import Font
from pymongo import MongoClient
import tiktoken
from functools import lru_cache
import threading
from datetime import datetime

# Add this at the top of your utils.py after imports
import sys

# Configure logging to handle Unicode
if sys.platform == "win32":
    # Windows-specific logging configuration
    import codecs
    if sys.stdout.encoding != 'UTF-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('app.log')
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION")
AZURE_OPENAI_CHAT_DEPLOYMENT = os.getenv("AZURE_OPENAI_CHAT_DEPLOYMENT")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")
COSMOS_URI = os.getenv("COSMOS_URI")
COSMOS_DB = os.getenv("COSMOS_DB")
COSMOS_COLLECTION = os.getenv("COSMOS_COLLECTION")

MAX_INPUT_TOKENS = 200000
TPM_LIMIT = 245000
TPM_THRESHOLD = 0.6  # Reduced from 0.8 to be more conservative

user_token_stats = {}  # Key: username, Value: token statistics
token_lock = threading.Lock()

# Global token stats for current session (updated directly)
token_stats = {
    "llm_input_tokens": 0,
    "llm_output_tokens": 0,
    "embedding_tokens": 0,
    "llm_calls": 0,
    "embedding_calls": 0,
    "start_time": time.time()
}

CHUNK_TYPES = {"TEXT": "text", "IMAGE": "image"}

# Initialize Azure OpenAI with better error handling
def initialize_azure_openai():
    try:
        # Clean endpoint URL
        endpoint = AZURE_OPENAI_ENDPOINT.strip()
        if endpoint.endswith('/'):
            endpoint = endpoint[:-1]
        
        logger.info(f"Initializing Azure OpenAI with endpoint: {endpoint}")
        logger.info(f"Using chat deployment: {AZURE_OPENAI_CHAT_DEPLOYMENT}")
        
        llm = AzureChatOpenAI(
            openai_api_version=AZURE_OPENAI_API_VERSION,
            azure_deployment=AZURE_OPENAI_CHAT_DEPLOYMENT,
            azure_endpoint=endpoint,
            api_key=AZURE_OPENAI_API_KEY,
            temperature=0.1,
            timeout=120,
            max_retries=2
        )
        
        vision_llm = AzureChatOpenAI(
            openai_api_version=AZURE_OPENAI_API_VERSION,
            azure_deployment="gpt-4o",
            azure_endpoint=endpoint,
            api_key=AZURE_OPENAI_API_KEY,
            temperature=0.1,
            timeout=120,
            max_retries=2
        )
        
        # Test the connection
        test_response = llm.invoke("Say 'connection test' only.")
        logger.info("Azure OpenAI connection test successful")
        
        return llm, vision_llm
    except Exception as e:
        logger.error(f"Failed to initialize Azure OpenAI: {str(e)}")
        raise

try:
    llm, vision_llm = initialize_azure_openai()
except Exception as e:
    logger.error(f"Azure OpenAI initialization failed: {e}")
    llm = None
    vision_llm = None

# Initialize Azure OpenAI Embeddings
try:
    endpoint = AZURE_OPENAI_ENDPOINT.strip()
    if endpoint.endswith('/'):
        endpoint = endpoint[:-1]
        
    embedding_model = AzureOpenAIEmbeddings(
        azure_deployment=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
        openai_api_version=AZURE_OPENAI_API_VERSION,
        azure_endpoint=endpoint,
        api_key=AZURE_OPENAI_API_KEY,
        model=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
        timeout=60
    )
    test_embedding = embedding_model.embed_query("test")
    logger.info("Azure OpenAI Embeddings connection test successful")
except Exception as e:
    logger.error(f"Failed to initialize Azure OpenAI Embeddings: {str(e)}")
    embedding_model = None

# Initialize Cosmos DB with better error handling
def initialize_cosmos_db():
    try:
        logger.info(f"Initializing Cosmos DB connection to: {COSMOS_DB}.{COSMOS_COLLECTION}")
        
        client = MongoClient(
            COSMOS_URI,
            connectTimeoutMS=30000,
            socketTimeoutMS=30000,
            retryWrites=True,
            appname="SOW-Analyzer"
        )
        db = client[COSMOS_DB]
        collection = db[COSMOS_COLLECTION]
        
        # Test connection with ping
        ping_result = client.admin.command('ping')
        logger.info(f"Cosmos DB connection successful: {ping_result}")
        
        # Test basic operations
        count = collection.count_documents({})
        logger.info(f"Cosmos DB has {count} existing documents")
        
        return client, db, collection
    except Exception as e:
        logger.error(f"Failed to connect to Cosmos DB: {str(e)}")
        # Log safe connection info
        if COSMOS_URI:
            safe_uri = COSMOS_URI.split('@')[-1] if '@' in COSMOS_URI else COSMOS_URI
            logger.error(f"Connection attempt to: {safe_uri}")
        raise

try:
    client, db, collection = initialize_cosmos_db()
except Exception as e:
    logger.error(f"Cosmos DB connection failed: {e}")
    client = None
    db = None
    collection = None

# Connection test functions for debug endpoint
def test_azure_openai_connection():
    """Test Azure OpenAI connection for debug endpoint"""
    try:
        if llm is None:
            return {"status": "failed", "error": "Azure OpenAI not initialized"}
        
        test_response = llm.invoke("Say 'connection test' only.")
        return {
            "status": "connected", 
            "test_response": test_response.content,
            "deployment": AZURE_OPENAI_CHAT_DEPLOYMENT
        }
    except Exception as e:
        return {
            "status": "failed", 
            "error": str(e),
            "deployment": AZURE_OPENAI_CHAT_DEPLOYMENT
        }

def test_cosmos_connection():
    """Test Cosmos DB connection for debug endpoint"""
    try:
        if collection is None:
            return {"status": "failed", "error": "Cosmos DB collection not initialized"}
        
        ping_result = client.admin.command('ping')
        count = collection.count_documents({})
        return {
            "status": "connected",
            "ping": "ok",
            "document_count": count,
            "database": COSMOS_DB,
            "collection": COSMOS_COLLECTION
        }
    except Exception as e:
        return {
            "status": "failed", 
            "error": str(e),
            "database": COSMOS_DB,
            "collection": COSMOS_COLLECTION
        }

@lru_cache(maxsize=50000)
def get_query_embedding(query: str):
    if embedding_model is None:
        raise Exception("Embedding model not initialized")
    return embedding_model.embed_query(query)

@lru_cache(maxsize=100000)
def cached_doc_embedding(text: str):
    if embedding_model is None:
        raise Exception("Embedding model not initialized")
    return embedding_model.embed_documents([text])[0]

def check_tpm_limit():
    """FIXED: Proper TPM limit checking"""
    current_time = time.time()
    elapsed_seconds = current_time - token_stats["start_time"]
    
    # Reset if we're in a new minute
    if elapsed_seconds >= 60:
        token_stats["llm_input_tokens"] = 0
        token_stats["llm_output_tokens"] = 0
        token_stats["embedding_tokens"] = 0
        token_stats["start_time"] = current_time
        logger.info("TPM counter reset - new minute started")
        return
    
    # Calculate CURRENT minute's TPM (not average)
    current_minute_tpm = token_stats["llm_input_tokens"] + token_stats["llm_output_tokens"]
    
    if current_minute_tpm > (TPM_LIMIT * TPM_THRESHOLD):
        sleep_time = 60 - elapsed_seconds
        logger.info(f"Approaching TPM limit ({current_minute_tpm:.0f}/{TPM_LIMIT}). Pausing for {sleep_time:.1f}s")
        time.sleep(sleep_time)
        # Reset for new minute
        token_stats["llm_input_tokens"] = 0
        token_stats["llm_output_tokens"] = 0
        token_stats["embedding_tokens"] = 0
        token_stats["start_time"] = time.time()

def count_tokens(text: str, model: str = "gpt-4o"):
    if not text or not isinstance(text, str):
        return 0
    try:
        encoding = tiktoken.encoding_for_model(model)
    except KeyError:
        encoding = tiktoken.get_encoding("cl100k_base")
    return len(encoding.encode(text))

def truncate_context(context, max_tokens=MAX_INPUT_TOKENS):
    if not context or not isinstance(context, str):
        return ""
    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(context)
    if len(tokens) <= max_tokens:
        return context
    keep_start = int(max_tokens * 0.7)
    keep_end = max_tokens - keep_start
    truncated_tokens = tokens[:keep_start] + tokens[-keep_end:]
    return encoding.decode(truncated_tokens)

def create_vector_index():
    try:
        if collection is None:
            logger.error("Collection is None, cannot create vector index")
            return False
            
        existing_indexes = list(collection.list_indexes())
        vector_index_exists = any(idx.get('name') == 'vectorSearchIndex' for idx in existing_indexes)
        if not vector_index_exists:
            index_definition = {
                "createIndexes": COSMOS_COLLECTION,
                "indexes": [
                    {
                        "name": "vectorSearchIndex",
                        "key": {"embedding": "cosmosSearch"},
                        "cosmosSearchOptions": {
                            "kind": "vector-ivf",
                            "numLists": 1,
                            "similarity": "COS",
                            "dimensions": 1536
                        }
                    }
                ]
            }
            db.command(index_definition)
        collection.create_index([("text_hash", 1)], name="text_hash_index", unique=False, sparse=True)
        collection.create_index([("desc_hash", 1)], name="desc_hash_index", unique=False, sparse=True)
        collection.create_index([("document_id", 1)], name="document_id_index")
        return True
    except Exception as e:
        logger.error(f"Index creation failed: {e}")
        return False

# FIXED: Use proper None check instead of truth value testing
vector_index_available = create_vector_index() if collection is not None else False

def extract_pdf_content_pymupdf(pdf_path):
    """Optimized PDF extraction"""
    text_content = ""
    images_content = []
    try:
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            
            # Extract text
            text = page.get_text()
            if text and text.strip():
                text_content += f"\n\n--- Page {page_num + 1} ---\n\n{text}"
            
        doc.close()
        return text_content, images_content
    except Exception as e:
        logger.error(f"Error extracting PDF content with PyMuPDF: {e}")
        return "", []    

def analyze_image_for_durations(image_b64):
    try:
        check_tpm_limit()
        prompt = """
        Analyze this image for project timeline or Gantt chart information. Look for:
        1. Phase durations (PREP, EXPLORE, REALIZE, DEPLOY, RUN phases)
        2. Sprint durations or counts
        3. Timeline bars showing months/weeks
        4. Any duration numbers or time spans
        If this appears to be a timeline/Gantt chart:
        - Count bar lengths or time spans
        - Sum sprint durations (e.g., 7 sprints × 3 weeks = 21 weeks)
        - Convert to months if needed (4 weeks ≈ 1 month)
        Output format: "Phase: Duration" for each phase found, or "No timeline data" if none found.
        Be concise and focus only on duration information.
        """
        input_tokens = count_tokens(prompt, model="gpt-4o")
       
        # Start timing for vision AI processing
        ai_start_time = time.time()
       
        token_stats["llm_input_tokens"] += input_tokens
        token_stats["llm_calls"] += 1
        messages = [
            HumanMessage(
                content=[
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
                ]
            )
        ]
        response = vision_llm(messages)
        output_tokens = count_tokens(response.content, model="gpt-4o")
        token_stats["llm_output_tokens"] += output_tokens
       
        # Calculate vision AI processing time
        ai_processing_time = time.time() - ai_start_time
       
        # FIX: Ensure we return a string description and processing time
        description = response.content if response.content else "No timeline data found"
        return description, ai_processing_time  # Return tuple consistently
       
    except Exception as e:
        logger.warning(f"Vision analysis failed: {e}")
        return "Image analysis failed", 0  # Return tuple consistently even on error
    
def generate_document_id(pdf_content):
    if not pdf_content:
        pdf_content = ""
    if not isinstance(pdf_content, str):
        pdf_content = str(pdf_content)
    normalized = normalize_and_clean_text(pdf_content)
    return hashlib.md5(normalized.encode('utf-8')).hexdigest()

def normalize_and_clean_text(text):
    if not text:
        return ""
    text = str(text).lower().strip()
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text

def check_existing_chunks(document_id):
    try:
        if collection is None:
            return False
        # Only check if any chunks exist, don't count all documents
        existing = collection.find_one(
            {"document_id": document_id}, 
            {"_id": 1}  # Only fetch _id for existence check
        )
        return existing is not None
    except Exception as e:
        logger.error(f"Error checking existing chunks: {e}")
        return False

def get_existing_embedding(text):
    try:
        if collection is None:
            return None
        text_hash = hashlib.md5(normalize_and_clean_text(text).encode('utf-8')).hexdigest()
        existing = collection.find_one({
            "text_hash": text_hash,
            "chunk_type": CHUNK_TYPES["TEXT"]
        })
        return existing["embedding"] if existing and "embedding" in existing else None
    except Exception as e:
        logger.error(f"Error getting existing embedding: {e}")
        return None

def store_chunks_in_cosmos(text_chunks, image_chunks, document_id):
    """Skip storing chunks to reduce latency but keep functionality"""
    logger.info("Skipping chunk storage for performance optimization")
    return True

def similarity_search_cosmos(query_text, document_id, k=5):
    """Skip vector search to reduce latency"""
    return []

# OPTIMIZED: Reduced template sizes
duration_template = PromptTemplate.from_template("""
Extract durations for PREP, EXPLORE, REALIZE, DEPLOY, RUN phases from:
{context}

Output JSON only:
{{"durations": {{"PREP": "", "EXPLORE": "", "REALIZE": "", "DEPLOY": "", "RUN": ""}}}}
""")

task_template = PromptTemplate.from_template("""
Check if these tasks are explicitly mentioned in the SOW:
{context}

Tasks to check:
{tasks_string}

Output JSON only:
{{"tasks": {{"task1": "yes/no", "task2": "yes/no", ...}}}}
""")

def safe_invoke(prompt, max_retries=2):
    if llm is None:
        return None, 0
        
    retries = 0
    while retries < max_retries:
        try:
            # Check TPM limit BEFORE processing
            check_tpm_limit()
            
            truncated_prompt = truncate_context(prompt, MAX_INPUT_TOKENS)
            input_tokens = count_tokens(truncated_prompt, model="gpt-4o")
            
            # Log token usage for debugging
            logger.info(f"Processing {input_tokens} input tokens. Current TPM: {token_stats['llm_input_tokens'] + token_stats['llm_output_tokens']}")
            
            token_stats["llm_input_tokens"] += input_tokens
            token_stats["llm_calls"] += 1
            
            # Start timing for AI processing
            ai_start_time = time.time()
            
            messages = [HumanMessage(content=truncated_prompt)]
            response_obj = llm(messages)
            output_text = response_obj.content
            output_tokens = count_tokens(output_text, model="gpt-4o")
            token_stats["llm_output_tokens"] += output_tokens
            
            # Calculate actual AI processing time
            ai_processing_time = time.time() - ai_start_time
            
            match = re.search(r'\{.*\}', output_text, re.DOTALL)
            if match:
                return json.loads(match.group()), ai_processing_time
        except Exception as e:
            logger.error(f"LLM Error (Attempt {retries+1}): {e}")
            retries += 1
    logger.error("Max retries reached for LLM invocation")
    return None, 0

def verify_substring_match(task, context):
    """ACCURATE substring matching with proper normalization"""
    if not task or not context:
        return "no"
    try:
        # Use proper normalization for accurate matching
        norm_task = normalize_and_clean_text(task)
        norm_context = normalize_and_clean_text(context)
        
        # Check for exact phrase matching with word boundaries for accuracy
        words = norm_task.split()
        if len(words) >= 3:  # For longer phrases, require more precision
            # Check if all major words are present in context
            major_words = [word for word in words if len(word) > 3]
            if major_words:
                all_major_present = all(word in norm_context for word in major_words)
                return "yes" if all_major_present else "no"
        
        # For shorter tasks, use exact substring but be more strict
        return "yes" if norm_task in norm_context else "no"
    except Exception as e:
        logger.error(f"Error in substring match: {e}")
        return "no"

def fuzzy_match_optimized(task, pdf_text, threshold=80, window=4000, step=3000):
    """ACCURATE fuzzy matching with higher threshold"""
    if not task or not pdf_text:
        return "no"
    try:
        task_str = normalize_and_clean_text(task)
        pdf_str = normalize_and_clean_text(pdf_text)
        best_score = 0
        
        # Use higher threshold for accuracy
        for start in range(0, len(pdf_str), step):
            snippet = pdf_str[start:start+window]
            score = fuzz.partial_ratio(task_str, snippet)
            if score > best_score:
                best_score = score
            if best_score >= threshold:  # Higher threshold for accuracy
                return "yes"
        return "yes" if best_score >= threshold else "no"
    except Exception as e:
        logger.error(f"Error in fuzzy_match_optimized: {e}")
        return "no"

def extract_durations_optimized(pdf_text):
    """Optimized but accurate duration extraction"""
    try:
        duration_context = truncate_context(pdf_text, 8000)  # Reduced from 12000
        prompt = duration_template.format(context=duration_context)
        response, ai_time = safe_invoke(prompt)
        if response and "durations" in response:
            return response["durations"], ai_time
    except Exception as e:
        logger.error(f"Error extracting durations: {e}")
    return {
        "PREP": "",
        "EXPLORE": "",
        "REALIZE": "",
        "DEPLOY": "",
        "RUN": ""
    }, 0

def process_batch_with_fallback_accurate(sub_batch, durations, normalized_pdf_text, pdf_text):
    """OPTIMIZED: Process batch with reduced context size"""
    try:
        batch_results = []
        total_ai_time = 0
        
        # Step 1: Check substring matches with proper normalization
        substring_flags = {}
        for heading, task in sub_batch:
            if not task:
                continue
            task_str = str(task)
            substring_flags[task_str] = verify_substring_match(task_str, normalized_pdf_text)
        
        # Step 2: Prepare tasks for AI processing - only those that need it
        tasks_for_ai_processing = []
        
        for i, (heading, task) in enumerate(sub_batch):
            if not task:
                continue
            task_str = str(task)
            
            # If clear substring match found, use that result
            if substring_flags.get(task_str) == "yes":
                batch_results.append({
                    "Heading": str(heading),
                    "Task": task_str,
                    "Present": "yes",
                    **durations
                })
            else:
                # Add to AI processing queue
                tasks_for_ai_processing.append((i, heading, task_str))
        
        # Step 3: Process remaining tasks with AI for accurate detection
        if tasks_for_ai_processing:
            # Use smaller context for AI processing
            context = truncate_context(normalized_pdf_text, 10000)  # Reduced from 15000
            tasks_string = "\n".join([f"task{idx+1}: {task}" for idx, (_, _, task) in enumerate(tasks_for_ai_processing)])
            
            if context and tasks_string:
                prompt = task_template.format(context=context, tasks_string=tasks_string)
                llm_response, ai_time = safe_invoke(prompt)
                total_ai_time += ai_time if ai_time else 0
                
                # Process AI results accurately
                if llm_response and "tasks" in llm_response:
                    task_values = list(llm_response["tasks"].values())
                    
                    for idx, (original_idx, heading, task_str) in enumerate(tasks_for_ai_processing):
                        final_answer = "no"
                        
                        # Use AI result if available and valid
                        if idx < len(task_values):
                            ai_answer = str(task_values[idx]).lower().strip()
                            if ai_answer == "yes":
                                final_answer = "yes"
                            else:
                                # If AI says no, use fuzzy matching as fallback
                                final_answer = fuzzy_match_optimized(task_str, pdf_text, threshold=80)
                        else:
                            # If no AI result, use fuzzy matching
                            final_answer = fuzzy_match_optimized(task_str, pdf_text, threshold=80)
                        
                        batch_results.append({
                            "Heading": str(heading),
                            "Task": task_str,
                            "Present": final_answer,
                            **durations
                        })
                else:
                    # If AI fails, use fuzzy matching for all
                    for idx, (original_idx, heading, task_str) in enumerate(tasks_for_ai_processing):
                        final_answer = fuzzy_match_optimized(task_str, pdf_text, threshold=80)
                        batch_results.append({
                            "Heading": str(heading),
                            "Task": task_str,
                            "Present": final_answer,
                            **durations
                        })
        
        return batch_results, total_ai_time
        
    except Exception as e:
        logger.error(f"Error processing batch: {e}")
        return [{"Heading": str(h), "Task": str(t), "Present": "error", **durations} for h, t in sub_batch], 0

def process_pdf_safely(uploaded_file):
    """Optimized but functional PDF processing"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf_file:
            tmp_pdf_file.write(uploaded_file.file.read() if hasattr(uploaded_file, "file") else uploaded_file.read())
            tmp_pdf_path = tmp_pdf_file.name
        
        try:
            pdf_text, images_content = extract_pdf_content_pymupdf(tmp_pdf_path)
            if not pdf_text or not pdf_text.strip():
                return None
            normalized_pdf_text = normalize_and_clean_text(pdf_text)
            if not normalized_pdf_text.strip():
                return None
            return pdf_text, normalized_pdf_text, tmp_pdf_path, images_content
        finally:
            # Ensure cleanup happens
            if tmp_pdf_path and os.path.exists(tmp_pdf_path):
                os.unlink(tmp_pdf_path)
    except Exception as e:
        logger.error(f"Error in process_pdf_safely: {str(e)}")
        return None

def create_excel_with_formatting(df, durations, output_file, activity_column_width=50):
    results_dict = {}
    for heading in df['Heading'].unique():
        tasks = df[df['Heading'] == heading]
        task_dict = {task: present for task, present in zip(tasks['Task'], tasks['Present'])}
        results_dict[heading] = task_dict
    yes_rows = []
    no_rows = []
    for category, tasks in results_dict.items():
        if category.upper() in durations:
            yes_rows.append({"Phase": category.upper(), "Duration": durations.get(category.upper(), ""), "Activity": ""})
        else:
            yes_rows.append({"Phase": "", "Duration": "", "Activity": category})
        for task, present in tasks.items():
            if present == "yes":
                yes_rows.append({"Phase": "", "Duration": "", "Activity": task})
        yes_rows.append({"Phase": "", "Duration": "", "Activity": ""})
        no_tasks = [task for task, present in tasks.items() if present == "no"]
        if no_tasks:
            if category.upper() in durations:
                no_rows.append({"Phase": category.upper(), "Duration": durations.get(category.upper(), ""), "Activity": ""})
            else:
                no_rows.append({"Phase": "", "Duration": "", "Activity": category})
            for task in no_tasks:
                no_rows.append({"Phase": "", "Duration": "", "Activity": task})
            no_rows.append({"Phase": "", "Duration": "", "Activity": ""})
    yes_df = pd.DataFrame(yes_rows)
    no_df = pd.DataFrame(no_rows)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if not yes_df.empty:
            yes_df.to_excel(writer, index=False, startrow=1, sheet_name="Yes_Tasks")
            workbook = writer.book
            worksheet = writer.sheets["Yes_Tasks"]
            activity_column_index = yes_df.columns.get_loc("Activity") + 1
            column_letter = get_column_letter(activity_column_index)
            worksheet.column_dimensions[column_letter].width = activity_column_width
            bold_font = Font(bold=True)
            for row in range(2, len(yes_rows) + 2):
                phase_cell = worksheet.cell(row=row, column=1)
                activity_cell = worksheet.cell(row=row, column=3)
                if phase_cell.value in durations.keys():
                    phase_cell.font = bold_font
                    worksheet.cell(row=row, column=2).font = bold_font
                if activity_cell.value in results_dict.keys() and activity_cell.value.upper() not in durations:
                    activity_cell.font = bold_font
            table = Table(displayName="PresentTasksTable", ref=worksheet.dimensions)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table.tableStyleInfo = style
            worksheet.add_table(table)
        if not no_df.empty:
            no_df.to_excel(writer, index=False, startrow=1, sheet_name="No_Tasks")
            worksheet_no = writer.sheets["No_Tasks"]
            activity_column_index = no_df.columns.get_loc("Activity") + 1
            column_letter = get_column_letter(activity_column_index)
            worksheet_no.column_dimensions[column_letter].width = activity_column_width
            for row in range(2, len(no_rows) + 2):
                phase_cell = worksheet_no.cell(row=row, column=1)
                activity_cell = worksheet_no.cell(row=row, column=3)
                if phase_cell.value in durations.keys():
                    phase_cell.font = bold_font
                    worksheet_no.cell(row=row, column=2).font = bold_font
                if activity_cell.value in results_dict.keys() and activity_cell.value.upper() not in durations:
                    activity_cell.font = bold_font
            table_no = Table(displayName="MissingTasksTable", ref=worksheet_no.dimensions)
            style_no = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table_no.tableStyleInfo = style_no
            worksheet_no.add_table(table_no)

def update_user_token_stats(local_stats, pdf_filename, username, email="Unknown", login_time="Unknown", logout_time="Not logged out", ai_response_time=0, total_processing_time=0):
    """Update user-specific token stats with data from a processing session"""
    with token_lock:
        # Initialize user stats if not exists
        if username not in user_token_stats:
            user_token_stats[username] = {
                "llm_input_tokens": 0,
                "llm_output_tokens": 0,
                "embedding_tokens": 0,
                "llm_calls": 0,
                "embedding_calls": 0,
                "total_files_processed": 0,
                "processing_sessions": [],
                "user_info": {
                    "username": username,
                    "email": email,
                    "first_login": login_time,
                    "last_activity": login_time
                }
            }
        
        user_stats = user_token_stats[username]
        
        # Update cumulative user stats
        user_stats["llm_input_tokens"] += local_stats.get("llm_input_tokens", 0)
        user_stats["llm_output_tokens"] += local_stats.get("llm_output_tokens", 0)
        user_stats["embedding_tokens"] += local_stats.get("embedding_tokens", 0)
        user_stats["llm_calls"] += local_stats.get("llm_calls", 0)
        user_stats["embedding_calls"] += local_stats.get("embedding_calls", 0)
        user_stats["total_files_processed"] += 1
        user_stats["user_info"]["last_activity"] = login_time
        
        # Calculate total tokens for this session
        session_total_tokens = (local_stats.get("llm_input_tokens", 0) + 
                              local_stats.get("llm_output_tokens", 0) + 
                              local_stats.get("embedding_tokens", 0))
        
        # Format total processing time for display
        if total_processing_time >= 60:
            minutes = int(total_processing_time // 60)
            seconds = int(total_processing_time % 60)
            formatted_total_time = f"{minutes} min {seconds} sec"
        else:
            formatted_total_time = f"{total_processing_time:.1f} sec"
        
        # Format AI response time for internal tracking
        if ai_response_time >= 60:
            minutes = int(ai_response_time // 60)
            seconds = int(ai_response_time % 60)
            formatted_ai_time = f"{minutes} min {seconds} sec"
        else:
            formatted_ai_time = f"{ai_response_time:.1f} sec"
        
        # Store individual session data
        session_data = {
            "pdf_filename": pdf_filename,
            "username": username,
            "email": email,
            "llm_input_tokens": local_stats.get("llm_input_tokens", 0),
            "llm_output_tokens": local_stats.get("llm_output_tokens", 0),
            "embedding_tokens": local_stats.get("embedding_tokens", 0),
            "total_tokens": session_total_tokens,
            "llm_calls": local_stats.get("llm_calls", 0),
            "embedding_calls": local_stats.get("embedding_calls", 0),
            "timestamp": time.time(),
            "processing_time": ai_response_time,  # Actual AI processing time in seconds
            "formatted_ai_time": formatted_ai_time,  # Human-readable format for AI time
            "total_processing_time": total_processing_time,  # Total processing time in seconds
            "formatted_total_time": formatted_total_time,  # Human-readable format for total time
            "login_time": login_time,
            "logout_time": logout_time,  # Will be updated when user logs out
            "ai_response_time": ai_response_time,
            "session_start": login_time
        }
        user_stats["processing_sessions"].append(session_data)
        
        # Reset token_stats for the next session
        token_stats["llm_input_tokens"] = 0
        token_stats["llm_output_tokens"] = 0
        token_stats["embedding_tokens"] = 0
        token_stats["llm_calls"] = 0
        token_stats["embedding_calls"] = 0
        token_stats["start_time"] = time.time()

def update_user_logout_time(username, logout_time):
    """Update logout time for the user's current session"""
    with token_lock:
        if username in user_token_stats:
            user_sessions = user_token_stats[username]['processing_sessions']
            if user_sessions:
                # Find sessions that don't have a proper logout time
                for session in reversed(user_sessions):  # Check from most recent
                    if session.get('logout_time') in ['Not logged out', '']:
                        session['logout_time'] = logout_time
                        logger.info(f"Updated logout time for user {username}: {logout_time}")
                        break

def log_user_token_usage(username=None):
    """Log token usage for specific user or all users"""
    with token_lock:
        if username:
            # Log specific user's usage
            if username in user_token_stats:
                user_stats = user_token_stats[username]
                logger.info(f"===== TOKEN USAGE REPORT FOR {username} =====")
                _log_user_stats(user_stats)
            else:
                logger.info(f"No token data found for user: {username}")
        else:
            # Log all users' usage
            logger.info("===== TOKEN USAGE REPORT - ALL USERS =====")
            for username, user_stats in user_token_stats.items():
                logger.info(f"--- User: {username} ---")
                _log_user_stats(user_stats)
                logger.info("")

def _log_user_stats(user_stats):
    """Helper function to log user statistics"""
    logger.info(f"Total Files Processed: {user_stats['total_files_processed']}")
    logger.info(f"Total LLM Calls: {user_stats['llm_calls']}")
    logger.info(f"Total Embedding Calls: {user_stats['embedding_calls']}")
    logger.info(f"Total LLM Input Tokens: {user_stats['llm_input_tokens']}")
    logger.info(f"Total LLM Output Tokens: {user_stats['llm_output_tokens']}")
    logger.info(f"Total Embedding Tokens: {user_stats['embedding_tokens']}")
    total_tokens = (user_stats['llm_input_tokens'] + 
                   user_stats['llm_output_tokens'] + 
                   user_stats['embedding_tokens'])
    logger.info(f"TOTAL Tokens: {total_tokens}")
    
    # Calculate average tokens per minute if we have session data
    if user_stats['processing_sessions']:
        total_processing_time = sum(session.get('processing_time', 0) for session in user_stats['processing_sessions'])
        if total_processing_time > 0:
            tpm = total_tokens / (total_processing_time / 60)
            logger.info(f"Average Tokens Per Minute: {tpm:.0f}")

def get_user_token_summary():
    """Get summary of token usage for all users"""
    with token_lock:
        summary = {}
        for username, user_stats in user_token_stats.items():
            total_tokens = (user_stats['llm_input_tokens'] + 
                           user_stats['llm_output_tokens'] + 
                           user_stats['embedding_tokens'])
            
            # Calculate total response time across all sessions
            total_response_time = 0
            if user_stats['processing_sessions']:
                for session in user_stats['processing_sessions']:
                    total_response_time += session.get('total_processing_time', 0)
            
            # Format total response time for display
            if total_response_time >= 60:
                minutes = int(total_response_time // 60)
                seconds = int(total_response_time % 60)
                formatted_total_time = f"{minutes} min {seconds} sec"
            else:
                formatted_total_time = f"{total_response_time:.1f} sec"
            
            # Get first and last activity from processing sessions
            first_login = "Unknown"
            last_activity = "Unknown"
            last_logout = "Not logged out"
            
            if user_stats['processing_sessions']:
                # Sort sessions by timestamp to find first and last
                sorted_sessions = sorted(user_stats['processing_sessions'], key=lambda x: x.get('timestamp', 0))
                first_login = sorted_sessions[0].get('login_time', 'Unknown') if sorted_sessions else "Unknown"
                last_activity = sorted_sessions[-1].get('login_time', 'Unknown') if sorted_sessions else "Unknown"
                last_logout = sorted_sessions[-1].get('logout_time', 'Not logged out') if sorted_sessions else "Not logged out"
            
            summary[username] = {
                "total_files_processed": user_stats['total_files_processed'],
                "total_tokens": total_tokens,
                "llm_calls": user_stats['llm_calls'],
                "total_response_time": formatted_total_time,  # Replaced embedding_calls with total_response_time
                "first_login": first_login,
                "last_activity": last_activity,
                "last_logout": last_logout
            }
        return summary
        
def create_token_report_excel(username=None):
    """Create an Excel file with comprehensive token usage statistics for specific user or all users"""
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        logger.info(f"Creating token report for user: {username}")
        
        with token_lock:
            session_rows = []
            
            if username:
                # Specific user report - get ALL sessions for this user
                if username not in user_token_stats:
                    logger.warning(f"No token data found for user: {username}")
                    return None
                user_sessions = user_token_stats[username]['processing_sessions']
                report_title = f"Token Usage Report - {username}"
                logger.info(f"Found {len(user_sessions)} sessions for user {username}")
            else:
                # All users report - get ALL sessions from ALL users
                user_sessions = []
                for current_user, user_data in user_token_stats.items():
                    user_sessions.extend(user_data.get('processing_sessions', []))
                
                report_title = "Token Usage Report - All Users"
                logger.info(f"Found {len(user_sessions)} total sessions across {len(user_token_stats)} users")
            
            # Include ALL sessions in the report
            for session in user_sessions:
                total_tokens = session.get('total_tokens', 0)
                
                session_rows.append({
                    "Username": session.get('username', 'Unknown'),
                    "File Name": session.get('pdf_filename', 'Unknown'),
                    "Total Tokens": total_tokens,
                    "Total Processing Time": session.get('formatted_total_time', '0 sec'),
                    "User Login Time": session.get('login_time', 'Unknown')
                })
        
        # Create DataFrame with ALL sessions
        sessions_df = pd.DataFrame(session_rows) if session_rows else pd.DataFrame(columns=[
            "Username", "File Name", "Total Tokens", "Total Processing Time", 
            "User Login Time"
        ])
        
        logger.info(f"Created DataFrame with {len(sessions_df)} rows")
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Individual sessions sheet with ALL data
                sessions_df.to_excel(writer, index=False, sheet_name='Token Usage Report')
                
                # Get workbook and worksheet
                workbook = writer.book
                sessions_ws = writer.sheets['Token Usage Report']
                
                # Add summary sheet for all users (existing)
                if not username:
                    summary_data = get_user_token_summary()
                    summary_rows = []
                    for user_name, stats in summary_data.items():
                        summary_rows.append({
                            "Username": user_name,
                            "Total Files Processed": stats['total_files_processed'],
                            "Total Tokens": stats['total_tokens'],
                            "Total Response Time": stats['total_response_time'],
                            "Last Activity": stats['last_activity']
                        })
                    
                    summary_df = pd.DataFrame(summary_rows)
                    summary_df.to_excel(writer, index=False, sheet_name='User Summary')
                    summary_ws = writer.sheets['User Summary']
                    
                    # Format summary sheet
                    header_fill = PatternFill(start_color="2867C5", end_color="2867C5", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in summary_ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Adjust column widths for summary sheet
                    summary_column_widths = [20, 15, 15, 20, 20]
                    for col_idx, width in enumerate(summary_column_widths, 1):
                        summary_ws.column_dimensions[get_column_letter(col_idx)].width = width
                
                # Apply formatting to sessions sheet
                header_fill = PatternFill(start_color="2867B5", end_color="2867B5", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in sessions_ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Adjust column widths for sessions sheet to accommodate all data
                session_column_widths = [20, 40, 15, 20, 20]
                for col_idx, width in enumerate(session_column_widths, 1):
                    sessions_ws.column_dimensions[get_column_letter(col_idx)].width = width
                
                # Reorder sheets for better organization (only for admin report)
                if not username:
                    # Order: User Summary, Token Usage Report
                    workbook._sheets.sort(key=lambda ws: {
                        'User Summary': 0,
                        'Token Usage Report': 1
                    }.get(ws.title, 2))
            
            output.seek(0)
            logger.info("Token report created successfully with all sessions")
            return output
            
        except Exception as excel_error:
            logger.error(f"Excel creation error: {str(excel_error)}")
            raise
        
    except Exception as e:
        logger.error(f"Error creating token report: {str(e)}", exc_info=True)
        return None


